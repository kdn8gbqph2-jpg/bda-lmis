import csv
import io
import logging
from io import BytesIO

from django.core.cache import cache
from django.db import transaction
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .filters import PlotFilter
from .models import Plot, PlotKhasraMapping
from .serializers import (
    PlotListSerializer,
    PlotDetailSerializer,
    PlotWriteSerializer,
    PlotGeoJSONSerializer,
)
from users.permissions import IsAdmin, IsStaffOrAbove

logger = logging.getLogger(__name__)

# ── Plot bulk-import template schema ─────────────────────────────────────────
# Columns the user fills in; khasra_number is resolved server-side to a
# Khasra FK using (colony, number).

_TEMPLATE_COLUMNS = ['plot_number', 'khasra_number', 'type', 'area_sqy', 'status']
_TEMPLATE_HINTS = [
    'e.g. 1A',          # plot_number
    'e.g. 1448',        # khasra_number
    'Residential / Commercial',
    'Square yards',
    'available / patta_ok / patta_missing / cancelled',
]


class PlotViewSet(viewsets.ModelViewSet):
    """
    GET    /api/plots/              list  (authenticated)
    POST   /api/plots/              create (staff+)
    GET    /api/plots/{id}/         detail
    PUT    /api/plots/{id}/         update (staff+)
    DELETE /api/plots/{id}/         soft-delete → status='cancelled' (admin)
    GET    /api/plots/{id}/pattas/  pattas covering this plot
    GET    /api/plots/{id}/documents/
    GET    /api/plots/{id}/history/
    POST   /api/plots/bulk-import/  CSV upload (admin)
    GET    /api/plots/geojson/      FeatureCollection (?colony=1&status=patta_ok)
    """

    queryset         = Plot.objects.select_related(
        'colony', 'primary_khasra', 'updated_by'
    ).prefetch_related(
        'khasra_mappings__khasra',
        'patta_mappings__patta',
    ).all()
    filterset_class  = PlotFilter
    search_fields    = ['plot_number']
    ordering_fields  = ['plot_number', 'area_sqm', 'status', 'type']
    ordering         = ['colony', 'plot_number']

    def get_permissions(self):
        if self.action == 'destroy':
            return [IsAuthenticated(), IsAdmin()]
        if self.action in ('create', 'update', 'partial_update',
                           'bulk_import', 'bulk_import_xlsx'):
            return [IsAuthenticated(), IsStaffOrAbove()]
        return [IsAuthenticated()]

    def get_serializer_class(self):
        if self.action == 'list':
            return PlotListSerializer
        if self.action in ('create', 'update', 'partial_update'):
            return PlotWriteSerializer
        return PlotDetailSerializer

    # ── Soft-delete: set status to cancelled ─────────────────────────────────

    def destroy(self, request, *args, **kwargs):
        plot = self.get_object()
        plot.status     = 'cancelled'
        plot.updated_by = request.user
        plot.save(update_fields=['status', 'updated_by', 'updated_at'])
        self._bust_geojson_cache(plot.colony_id)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_update(self, serializer):
        instance = serializer.save(updated_by=self.request.user)
        self._bust_geojson_cache(instance.colony_id)

    def perform_create(self, serializer):
        instance = serializer.save(updated_by=self.request.user)
        self._bust_geojson_cache(instance.colony_id)

    @staticmethod
    def _bust_geojson_cache(colony_id):
        cache.delete(f'plots:geojson:colony:{colony_id}')
        cache.delete('plots:geojson:all')

    # ── /api/plots/{id}/pattas/ ───────────────────────────────────────────────

    @action(detail=True, methods=['get'])
    def pattas(self, request, pk=None):
        plot = self.get_object()
        try:
            from pattas.models import Patta, PlotPattaMapping
            from pattas.serializers import PattaListSerializer
            patta_ids = PlotPattaMapping.objects.filter(
                plot=plot
            ).values_list('patta_id', flat=True)
            pattas = Patta.objects.filter(id__in=patta_ids)
            return Response(PattaListSerializer(pattas, many=True).data)
        except Exception:
            return Response(
                {'detail': 'Pattas app not yet available.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

    # ── /api/plots/{id}/documents/ ────────────────────────────────────────────

    @action(detail=True, methods=['get'])
    def documents(self, request, pk=None):
        plot = self.get_object()
        try:
            from documents.models import Document
            from documents.serializers import DocumentListSerializer
            docs = Document.objects.filter(linked_plot=plot)
            return Response(DocumentListSerializer(docs, many=True).data)
        except Exception:
            return Response(
                {'detail': 'Documents app not yet available.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

    # ── /api/plots/{id}/history/ ──────────────────────────────────────────────

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        plot = self.get_object()
        try:
            from audit.models import AuditLog
            logs = AuditLog.objects.filter(
                entity_type='plot', entity_id=plot.id
            ).order_by('-timestamp')[:50]
            from audit.serializers import AuditLogSerializer
            return Response(AuditLogSerializer(logs, many=True).data)
        except Exception:
            return Response([])

    # ── /api/plots/bulk-import/ ───────────────────────────────────────────────

    @action(
        detail=False, methods=['post'],
        url_path='bulk-import',
        parser_classes=[MultiPartParser, FormParser],
    )
    def bulk_import(self, request):
        """
        Accept a CSV file with columns:
        plot_number, colony_id, primary_khasra_id, type, area_sqy, status

        Returns: {created: N, updated: N, errors: [...]}
        """
        csv_file = request.FILES.get('file')
        if not csv_file:
            return Response(
                {'detail': 'No file provided. Send multipart field "file".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            decoded = csv_file.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            return Response(
                {'detail': 'File must be UTF-8 encoded.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        reader  = csv.DictReader(io.StringIO(decoded))
        created = 0
        updated = 0
        errors  = []

        required_cols = {'plot_number', 'colony_id', 'primary_khasra_id', 'type', 'area_sqy'}
        if not required_cols.issubset(set(reader.fieldnames or [])):
            missing = required_cols - set(reader.fieldnames or [])
            return Response(
                {'detail': f'CSV missing required columns: {missing}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            for i, row in enumerate(reader, start=2):   # row 1 = header
                try:
                    ser = PlotWriteSerializer(data={
                        'plot_number':       row['plot_number'].strip(),
                        'colony':            int(row['colony_id']),
                        'primary_khasra':    int(row['primary_khasra_id']),
                        'type':              row.get('type', 'Residential').strip(),
                        'area_sqy':          row['area_sqy'].strip() or None,
                        'status':            row.get('status', 'available').strip(),
                    })

                    existing = Plot.objects.filter(
                        plot_number=row['plot_number'].strip()
                    ).first()

                    if existing:
                        ser = PlotWriteSerializer(existing, data={
                            'plot_number':    row['plot_number'].strip(),
                            'colony':         int(row['colony_id']),
                            'primary_khasra': int(row['primary_khasra_id']),
                            'type':           row.get('type', 'Residential').strip(),
                            'area_sqy':       row['area_sqy'].strip() or None,
                            'status':         row.get('status', 'available').strip(),
                        })
                        if ser.is_valid():
                            ser.save(updated_by=request.user)
                            updated += 1
                        else:
                            errors.append({'row': i, 'errors': ser.errors})
                    else:
                        if ser.is_valid():
                            ser.save(updated_by=request.user)
                            created += 1
                        else:
                            errors.append({'row': i, 'errors': ser.errors})

                except Exception as exc:
                    errors.append({'row': i, 'errors': str(exc)})

        return Response({'created': created, 'updated': updated, 'errors': errors})

    # ── /api/plots/template/  (blank xlsx for offline data entry) ────────────

    @action(detail=False, methods=['get'])
    def template(self, request):
        """
        GET /api/plots/template/

        Streams a blank .xlsx whose header row matches the columns the
        bulk-import-xlsx action expects. Row 2 contains lightweight hints
        so the user knows what each cell should hold.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Plots'

        # Header
        ws.append(_TEMPLATE_COLUMNS)
        for col_idx in range(1, len(_TEMPLATE_COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = PatternFill('solid', fgColor='1E3A8A')
            cell.alignment = Alignment(horizontal='center')

        # Hint row (italic, light grey) — users delete or overwrite it
        ws.append(_TEMPLATE_HINTS)
        for col_idx in range(1, len(_TEMPLATE_HINTS) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = Font(italic=True, color='94A3B8')

        widths = [14, 16, 22, 14, 36]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
        ws.freeze_panes = 'A3'

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="plots_template.xlsx"'
        return resp

    # ── /api/plots/bulk-import-xlsx/  (filled template upload) ────────────────

    @action(
        detail=False, methods=['post'],
        url_path='bulk-import-xlsx',
        parser_classes=[MultiPartParser, FormParser],
    )
    def bulk_import_xlsx(self, request):
        """
        POST /api/plots/bulk-import-xlsx/
          multipart/form-data:
            file:   <plots_template.xlsx>
            colony: <colony_id>

        For each data row, look up Khasra by (colony, khasra_number); if it
        doesn't exist, create it with just the number (geometry can be added
        later). Plot is upserted by plot_number.
        """
        import openpyxl
        from colonies.models import Colony, Khasra

        xlsx_file = request.FILES.get('file')
        colony_id = request.data.get('colony')

        if not xlsx_file:
            return Response({'detail': 'Send multipart field "file".'},
                            status=status.HTTP_400_BAD_REQUEST)
        if not colony_id:
            return Response({'detail': 'Send multipart field "colony" (colony id).'},
                            status=status.HTTP_400_BAD_REQUEST)
        try:
            colony = Colony.objects.get(pk=int(colony_id))
        except (Colony.DoesNotExist, ValueError, TypeError):
            return Response({'detail': 'Colony not found.'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        except Exception as exc:
            return Response({'detail': f'Unable to read workbook: {exc}'},
                            status=status.HTTP_400_BAD_REQUEST)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response({'detail': 'Workbook has no data rows.'},
                            status=status.HTTP_400_BAD_REQUEST)

        headers = [str(h).strip() if h else '' for h in rows[0]]
        # Skip an optional "hint" row that uses italics — treat as data only
        # if the plot_number cell parses as a sensible value.
        data_rows = rows[1:]
        if data_rows and (data_rows[0][0] is None or str(data_rows[0][0]).startswith('e.g.')):
            data_rows = data_rows[1:]

        col_index = {h: i for i, h in enumerate(headers) if h}
        required = {'plot_number', 'khasra_number'}
        missing  = required - set(col_index.keys())
        if missing:
            return Response(
                {'detail': f'Missing required columns: {", ".join(sorted(missing))}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def _cell(row, col):
            i = col_index.get(col)
            if i is None or i >= len(row): return None
            v = row[i]
            if v is None: return None
            if isinstance(v, float) and v.is_integer(): return str(int(v))
            return str(v).strip()

        created, updated, errors = 0, 0, []
        with transaction.atomic():
            for idx, row in enumerate(data_rows, start=2):
                if not any(row): continue
                plot_number = _cell(row, 'plot_number')
                khasra_num  = _cell(row, 'khasra_number')
                if not plot_number or not khasra_num:
                    continue   # silently skip empty rows
                khasra, _ = Khasra.objects.get_or_create(colony=colony, number=khasra_num)
                payload = {
                    'plot_number':    plot_number,
                    'colony':         colony.pk,
                    'primary_khasra': khasra.pk,
                    'type':           _cell(row, 'type')     or 'Residential',
                    'area_sqy':       _cell(row, 'area_sqy') or None,
                    'status':         _cell(row, 'status')   or 'available',
                }
                existing = Plot.objects.filter(plot_number=plot_number).first()
                ser = (PlotWriteSerializer(existing, data=payload)
                       if existing else PlotWriteSerializer(data=payload))
                if not ser.is_valid():
                    errors.append({'row': idx, 'errors': ser.errors})
                    continue
                ser.save(updated_by=request.user)
                if existing: updated += 1
                else:        created += 1

        logger.info('Plot xlsx import for colony %s by %s: created=%d updated=%d errors=%d',
                    colony.pk, request.user, created, updated, len(errors))
        self._bust_geojson_cache(colony.pk)
        return Response({'created': created, 'updated': updated, 'errors': errors})

    # ── /api/plots/geojson/ ───────────────────────────────────────────────────

    @action(detail=False, methods=['get'], url_path='geojson')
    def geojson_all(self, request):
        """
        FeatureCollection of all plots.  Supports ?colony=1 and ?status=patta_ok
        via the standard filter.  Results cached per colony for 30 min.
        """
        colony_id = request.query_params.get('colony')
        st        = request.query_params.get('status')

        # Simple cache key (colony + status combo)
        cache_key = f'plots:geojson:colony:{colony_id or "all"}:status:{st or "all"}'
        cached    = cache.get(cache_key)
        if cached:
            return Response(cached)

        qs   = self.filter_queryset(
            self.get_queryset().select_related('colony')
        )
        data = PlotGeoJSONSerializer.collection(qs)
        cache.set(cache_key, data, 60 * 30)   # 30 min TTL
        return Response(data)

    # ── /api/plots/{id}/geojson/ ──────────────────────────────────────────────

    @action(detail=True, methods=['get'], url_path='geojson')
    def geojson(self, request, pk=None):
        plot = self.get_object()
        return Response(PlotGeoJSONSerializer.feature(plot))
