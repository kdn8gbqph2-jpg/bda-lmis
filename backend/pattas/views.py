import logging
from io import BytesIO

from django.core.cache import cache
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .filters import PattaFilter
from .models import Patta, PlotPattaMapping, PattaVersion
from .serializers import (
    PattaListSerializer,
    PattaDetailSerializer,
    PattaWriteSerializer,
    PattaVersionSerializer,
)
from users.permissions import IsAdmin, IsStaffOrAbove
from approvals.mixins import StaffApprovalMixin

logger = logging.getLogger(__name__)


class PattaViewSet(StaffApprovalMixin, viewsets.ModelViewSet):
    """
    GET    /api/pattas/                     list   (authenticated)
    POST   /api/pattas/                     create (staff+)
    GET    /api/pattas/{id}/                detail
    PUT    /api/pattas/{id}/                update (staff+)
    DELETE /api/pattas/{id}/                soft-delete → status='cancelled' (admin)
    GET    /api/pattas/{id}/versions/       version history
    POST   /api/pattas/{id}/link-document/  attach a document (staff+)
    GET    /api/pattas/{id}/plots/          all plots covered by this patta
    """

    queryset        = Patta.objects.select_related(
        'colony', 'document', 'superseded_by', 'updated_by'
    ).prefetch_related('plot_mappings__plot').all()
    filterset_class = PattaFilter
    search_fields   = ['patta_number', 'allottee_name']
    ordering_fields = ['patta_number', 'allottee_name', 'issue_date', 'status']
    ordering        = ['colony', 'patta_number']

    # ── Staff approval gate ─────────────────────────────────────────────────
    # Staff JSON writes get queued as ChangeRequest rows; admin/super
    # writes (and any multipart submission) pass through directly.
    approval_target_type        = 'patta'
    approval_target_label_field = 'patta_number'

    def get_permissions(self):
        if self.action == 'destroy':
            return [IsAuthenticated(), IsAdmin()]
        if self.action in ('create', 'update', 'partial_update', 'link_document'):
            return [IsAuthenticated(), IsStaffOrAbove()]
        return [IsAuthenticated()]

    def get_serializer_class(self):
        if self.action == 'list':
            return PattaListSerializer
        if self.action in ('create', 'update', 'partial_update'):
            return PattaWriteSerializer
        return PattaDetailSerializer

    # ── Soft-delete ───────────────────────────────────────────────────────────

    def destroy(self, request, *args, **kwargs):
        patta = self.get_object()
        patta.status     = 'cancelled'
        patta.updated_by = request.user
        patta.save(update_fields=['status', 'updated_by', 'updated_at'])
        self._snapshot(patta, request.user)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_update(self, serializer):
        instance = serializer.save(updated_by=self.request.user)
        self._snapshot(instance, self.request.user)
        cache.delete(f'colony:{instance.colony_id}:stats')

    def perform_create(self, serializer):
        instance = serializer.save(updated_by=self.request.user)
        self._snapshot(instance, self.request.user)
        cache.delete(f'colony:{instance.colony_id}:stats')

    @staticmethod
    def _snapshot(patta, user):
        """Store a PattaVersion snapshot after every mutation."""
        import json
        from rest_framework.renderers import JSONRenderer
        snapshot = PattaDetailSerializer(patta).data
        try:
            raw = json.loads(JSONRenderer().render(snapshot))
        except Exception:
            raw = {}
        PattaVersion.objects.create(patta=patta, snapshot=raw, changed_by=user)

    # ── /api/pattas/export/ ───────────────────────────────────────────────────

    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        GET /api/pattas/export/?<same filter params as list>

        Streams an .xlsx of the currently filtered queryset. Authentication
        required (uses the same JWT as the rest of the API).
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        qs = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Patta Ledger'

        headers = [
            'Patta No.', 'Allottee Name', 'Allottee Address', 'Colony',
            'Plot(s)', 'Issue Date', 'Amendment Date',
            'Challan No.', 'Challan Date',
            'Lease Amount (₹)', 'Lease Duration',
            'Regulation File', 'Status', 'Remarks',
        ]
        ws.append(headers)
        header_font  = Font(bold=True, color='FFFFFF')
        header_fill  = PatternFill('solid', fgColor='1E3A8A')   # blue-900
        center_align = Alignment(horizontal='center', vertical='center')
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align

        for p in qs:
            plots = ', '.join(
                pm.plot.plot_number for pm in p.plot_mappings.all() if pm.plot
            )
            reg = ''
            if p.regulation_file_present is True:  reg = 'हाँ'
            elif p.regulation_file_present is False: reg = 'नही'
            ws.append([
                p.patta_number,
                p.allottee_name,
                p.allottee_address or '',
                p.colony.name if p.colony_id else '',
                plots,
                p.issue_date.isoformat()     if p.issue_date     else '',
                p.amendment_date.isoformat() if p.amendment_date else '',
                p.challan_number or '',
                p.challan_date.isoformat() if p.challan_date else '',
                float(p.lease_amount) if p.lease_amount is not None else '',
                p.lease_duration or '',
                reg,
                p.get_status_display() if hasattr(p, 'get_status_display') else p.status,
                p.remarks or '',
            ])

        # Reasonable column widths
        widths = [12, 28, 32, 28, 18, 12, 14, 16, 12, 14, 14, 12, 12, 30]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
        ws.freeze_panes = 'A2'

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        logger.info('Patta export: %d rows by user %s', qs.count(), request.user)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="patta_ledger.xlsx"'
        return resp

    # ── /api/pattas/{id}/versions/ ────────────────────────────────────────────

    @action(detail=True, methods=['get'])
    def versions(self, request, pk=None):
        patta    = self.get_object()
        versions = patta.versions.select_related('changed_by').order_by('-changed_at')[:20]
        return Response(PattaVersionSerializer(versions, many=True).data)

    # ── /api/pattas/{id}/link-document/ ──────────────────────────────────────

    @action(detail=True, methods=['post'], url_path='link-document')
    def link_document(self, request, pk=None):
        patta       = self.get_object()
        document_id = request.data.get('document_id')
        if not document_id:
            return Response(
                {'detail': 'document_id is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            from documents.models import Document
            doc = Document.objects.get(pk=document_id)
        except Exception:
            return Response(
                {'detail': 'Document not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        patta.document   = doc
        patta.updated_by = request.user
        patta.save(update_fields=['document', 'updated_by', 'updated_at'])
        return Response(PattaDetailSerializer(patta).data)

    # ── /api/pattas/{id}/plots/ ───────────────────────────────────────────────

    @action(detail=True, methods=['get'])
    def plots(self, request, pk=None):
        patta = self.get_object()
        try:
            from plots.models import Plot
            from plots.serializers import PlotListSerializer
            plot_ids = patta.plot_mappings.values_list('plot_id', flat=True)
            plots    = Plot.objects.filter(id__in=plot_ids).select_related(
                'colony', 'primary_khasra'
            )
            return Response(PlotListSerializer(plots, many=True).data)
        except Exception:
            return Response(
                {'detail': 'Plots app not yet available.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )
