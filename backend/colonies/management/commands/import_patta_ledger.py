"""
Management command: import_patta_ledger
=======================================

Usage:
  docker compose exec backend python manage.py import_patta_ledger \
      --file /path/to/Patta\ Ledger\ Format.xlsx \
      [--dry-run] [--colony "बौद्ध बिहार"]

Reads the BDA patta ledger Excel file.  Each sheet = one colony.

Sheet layout (per colony):
  Row 1: Colony name (योजना का नाम)
  Row 2: Village/Town name (ignored)
  Row 3: Chak number (चक नम्बर)
  Row 4: Layout approval date (लेआउट प्लान अनुमोदन दिनांक)
  Row 5: Total plots per layout plan (used for validation only)
  Row 6: Khasra numbers comma-separated (खसरा नम्बर)
  Row 7: Column headers (skipped)
  Row 8: Second header row (skipped)
  Row 9+: Data rows

Data columns (0-indexed after reading into openpyxl):
  A: Serial number (skip)
  B: Allottee name
  C: Allottee address
  D: Khasra number(s) — may be comma-separated
  E: Plot number
  F: Area in square yards
  G: Patta number (plain integer stored as string)
  H: Patta issue date
  I: Challan number
  J: Challan date
  K: Lease amount
  L: Lease duration (text)
  M: Regulation file present (हाँ / नही / blank)
  N: DMS file number (BHR102703 format, or "NO")
  O: Remarks
"""

import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


# ── helpers ───────────────────────────────────────────────────────────────────

def _str(val) -> str:
    """Coerce any cell value to a clean string."""
    if val is None:
        return ''
    return str(val).strip()


def _alnum(val) -> str:
    """
    Coerce a cell to a string suitable for ID-style fields (patta_number,
    plot_number, khasra number, challan number, DMS number). These can be
    alphanumeric ("17A", "BHR102703") but never decimal — Excel often stores
    plain integers as floats, producing "3498.0" / "1.0" when stringified.
    Strip the trailing ".0" but preserve genuine alphanumeric tokens.
    """
    if val is None:
        return ''
    # Treat int-valued floats as ints to avoid "3498.0"
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    s = str(val).strip()
    # Catch already-stringified ints like "3498.0"
    if re.fullmatch(r'-?\d+\.0+', s):
        return s.split('.')[0]
    return s


def _int(val) -> int | None:
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return None


def _decimal(val) -> Decimal | None:
    try:
        cleaned = re.sub(r'[^\d.]', '', str(val).strip())
        return Decimal(cleaned) if cleaned else None
    except InvalidOperation:
        return None


def _date(val):
    """Parse various date representations from Excel."""
    if val is None:
        return None
    if hasattr(val, 'date'):         # datetime object from openpyxl
        return val.date()
    if isinstance(val, str):
        val = val.strip()
        for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%m-%y'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _regulation(val) -> bool | None:
    s = _str(val).lower()
    if s in ('हाँ', 'han', 'yes', 'y', 'ha', 'हा', '1', 'true'):
        return True
    if s in ('नही', 'nahi', 'no', 'n', '0', 'false', 'नहीं'):
        return False
    return None


def _parse_khasra_numbers(raw: str) -> list[str]:
    """
    Split a raw khasra cell into individual khasra number strings.
    Handles: "1448,1449" / "1448, 1449" / "1448/1887" (single compound)
    A compound like "1450/1887" is ONE khasra number, not two.
    The separator between khasras is comma (or semicolon).
    """
    if not raw:
        return []
    parts = []
    for p in re.split(r'[,;]', raw):
        n = _alnum(p)
        if n:
            parts.append(n)
    return parts


# ── command ───────────────────────────────────────────────────────────────────

class Command(BaseCommand):
    help = 'Import patta ledger data from BDA Excel workbook into the database.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', required=True,
            help='Absolute path to the Excel file (Patta Ledger Format.xlsx)',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Parse and validate only; do not write to the database.',
        )
        parser.add_argument(
            '--colony', default=None,
            help='Import only the sheet with this colony name (Hindi). '
                 'Leave blank to import all sheets.',
        )

    def handle(self, *args, **options):
        import openpyxl

        file_path = options['file']
        dry_run   = options['dry_run']
        only_col  = options.get('colony')

        self.stdout.write(f'Opening: {file_path}')
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {file_path}')
        except Exception as e:
            raise CommandError(f'Cannot open workbook: {e}')

        totals = {'colonies': 0, 'khasras': 0, 'plots': 0,
                  'pattas': 0, 'mappings': 0, 'documents': 0,
                  'errors': 0}

        for sheet_name in wb.sheetnames:
            if only_col and sheet_name != only_col:
                continue
            # Skip Excel's default "Sheet1" / "Sheet2" placeholder tabs.
            if re.match(r'^Sheet\d+$', sheet_name, flags=re.IGNORECASE):
                self.stdout.write(f'  Skipping placeholder sheet: {sheet_name}')
                continue
            # Skip sheets with fewer than 9 rows (no data rows possible)
            ws_check = wb[sheet_name]
            if ws_check.max_row < 9:
                self.stdout.write(f'  Skipping blank sheet: {sheet_name}')
                continue

            ws = ws_check
            self.stdout.write(f'\n── Sheet: {sheet_name}')

            try:
                with transaction.atomic():
                    counts = self._import_sheet(ws, sheet_name, dry_run)
                    for k, v in counts.items():
                        totals[k] += v
                    if dry_run:
                        raise _DryRunRollback()
            except _DryRunRollback:
                self.stdout.write(self.style.WARNING('  [dry-run] rolled back'))
            except Exception as exc:
                self.stderr.write(f'  ERROR on sheet "{sheet_name}": {exc}')
                totals['errors'] += 1

        self.stdout.write('\n' + '─' * 60)
        self.stdout.write(self.style.SUCCESS(
            f'Done.  colonies={totals["colonies"]}  khasras={totals["khasras"]}  '
            f'plots={totals["plots"]}  pattas={totals["pattas"]}  '
            f'mappings={totals["mappings"]}  documents={totals["documents"]}  '
            f'sheet_errors={totals["errors"]}'
        ))
        if dry_run:
            self.stdout.write(self.style.WARNING('(dry-run — no data committed)'))

    # ── per-sheet import ──────────────────────────────────────────────────────

    def _import_sheet(self, ws, sheet_name: str, dry_run: bool) -> dict:
        from colonies.models  import Colony, Khasra
        from plots.models     import Plot, PlotKhasraMapping
        from pattas.models    import Patta, PlotPattaMapping
        from documents.models import Document
        from users.models     import CustomUser

        # Use the first superuser as the system uploader; fall back to None
        # (nullable FK) if no user exists yet.
        system_user = CustomUser.objects.filter(is_superuser=True).first()
        system_user_id = system_user.pk if system_user else None

        counts = {'colonies': 0, 'khasras': 0, 'plots': 0,
                  'pattas': 0, 'mappings': 0, 'documents': 0}

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            self.stdout.write(f'  Skipping — fewer than 6 rows')
            return counts

        # ── Parse header rows ──────────────────────────────────────────
        # Colony name: try col B first, fall back to sheet name.
        # Chak/date/khasras: the actual values are in col D (index 3).
        colony_name          = _str(rows[0][1]) or sheet_name   # row 1, col B or sheet name
        chak_number          = _int(rows[2][3])                  # row 3, col D
        layout_approval_date = _date(rows[3][3])                 # row 4, col D
        khasra_raw           = _str(rows[5][3])                  # row 6, col D

        self.stdout.write(
            f'  Colony: {colony_name} | Chak: {chak_number} | '
            f'Layout date: {layout_approval_date} | Khasras: {khasra_raw[:60]}'
        )

        # ── Upsert Colony ──────────────────────────────────────────────
        colony, created = Colony.objects.get_or_create(
            name=colony_name,
            defaults={
                'chak_number':          chak_number,
                'layout_approval_date': layout_approval_date,
            },
        )
        if not created:
            # Update date/chak if they were missing
            updated = False
            if layout_approval_date and not colony.layout_approval_date:
                colony.layout_approval_date = layout_approval_date
                updated = True
            if chak_number and not colony.chak_number:
                colony.chak_number = chak_number
                updated = True
            if updated:
                colony.save()
        counts['colonies'] += 1

        # ── Upsert Khasras from header row 6 ──────────────────────────
        khasra_numbers = _parse_khasra_numbers(khasra_raw)
        khasra_map: dict[str, Khasra] = {}  # number → Khasra instance

        for knum in khasra_numbers:
            khasra, _ = Khasra.objects.get_or_create(
                colony=colony, number=knum,
            )
            khasra_map[knum] = khasra
            counts['khasras'] += 1

        # ── Data rows start at row 9 (index 8) ────────────────────────
        # rows[6] = col headers (row 7), rows[7] = second header (row 8)
        data_rows = rows[8:]

        for row_idx, row in enumerate(data_rows, start=9):
            # Skip completely empty rows
            if all(v is None or _str(v) == '' for v in row):
                continue

            allottee_name = _str(row[1] if len(row) > 1 else None)
            if not allottee_name:
                continue   # skip rows without an allottee name

            allottee_address = _str(row[2] if len(row) > 2 else None)
            khasra_cell      = _str(row[3] if len(row) > 3 else None)
            plot_number      = _alnum(row[4] if len(row) > 4 else None)
            area_sqy         = _decimal(row[5] if len(row) > 5 else None)
            patta_number_raw = _alnum(row[6] if len(row) > 6 else None)
            issue_date       = _date(row[7] if len(row) > 7 else None)
            challan_number   = _alnum(row[8] if len(row) > 8 else None)
            challan_date     = _date(row[9] if len(row) > 9 else None)
            lease_amount     = _decimal(row[10] if len(row) > 10 else None)
            lease_duration   = _str(row[11] if len(row) > 11 else None)
            reg_file         = _regulation(row[12] if len(row) > 12 else None)
            dms_number       = _alnum(row[13] if len(row) > 13 else None)
            remarks          = _str(row[14] if len(row) > 14 else None)

            if not plot_number:
                self.stderr.write(f'  row {row_idx}: no plot number, skipping')
                continue

            # ── Resolve primary khasra for this plot ──────────────────
            row_khasra_nums = _parse_khasra_numbers(khasra_cell)
            primary_khasra  = None

            # Ensure all khasras from this row exist in the DB
            for knum in row_khasra_nums:
                if knum not in khasra_map:
                    k, _ = Khasra.objects.get_or_create(colony=colony, number=knum)
                    khasra_map[knum] = k
                    counts['khasras'] += 1

            if row_khasra_nums:
                primary_khasra = khasra_map[row_khasra_nums[0]]
            else:
                # Fall back to first colony khasra if none specified
                primary_khasra = (
                    list(khasra_map.values())[0]
                    if khasra_map else None
                )

            if primary_khasra is None:
                self.stderr.write(
                    f'  row {row_idx}: cannot determine khasra for plot {plot_number}, skipping'
                )
                continue

            # ── Upsert Plot ───────────────────────────────────────────
            plot, plot_created = Plot.objects.get_or_create(
                plot_number=plot_number,
                defaults={
                    'colony':          colony,
                    'primary_khasra':  primary_khasra,
                    'area_sqy':        area_sqy,
                    'status':          'patta_ok' if patta_number_raw else 'available',
                },
            )
            if not plot_created and area_sqy and not plot.area_sqy:
                plot.area_sqy = area_sqy
                plot.save()
            counts['plots'] += 1

            # ── PlotKhasraMapping for secondary khasras ───────────────
            for knum in row_khasra_nums[1:]:
                PlotKhasraMapping.objects.get_or_create(
                    plot=plot, khasra=khasra_map[knum],
                )
                counts['mappings'] += 1

            # ── Upsert Patta ──────────────────────────────────────────
            if not patta_number_raw:
                continue   # no patta for this plot yet

            patta_number = patta_number_raw.strip()

            patta, patta_created = Patta.objects.get_or_create(
                patta_number=patta_number,
                defaults={
                    'colony':                  colony,
                    'allottee_name':           allottee_name,
                    'allottee_address':        allottee_address,
                    'issue_date':              issue_date,
                    'challan_number':          challan_number,
                    'challan_date':            challan_date,
                    'lease_amount':            lease_amount,
                    'lease_duration':          lease_duration,
                    'regulation_file_present': reg_file,
                    'remarks':                 remarks,
                    'status':                  'issued',
                },
            )
            if not patta_created:
                # Fill in any blank fields from Excel
                changed = False
                for attr, val in [
                    ('allottee_address',        allottee_address),
                    ('challan_number',          challan_number),
                    ('challan_date',            challan_date),
                    ('lease_amount',            lease_amount),
                    ('lease_duration',          lease_duration),
                    ('regulation_file_present', reg_file),
                    ('remarks',                 remarks),
                ]:
                    if val not in (None, '') and not getattr(patta, attr):
                        setattr(patta, attr, val)
                        changed = True
                if changed:
                    patta.save()
            counts['pattas'] += 1

            # ── PlotPattaMapping ──────────────────────────────────────
            PlotPattaMapping.objects.get_or_create(
                plot=plot, patta=patta,
                defaults={'ownership_share_pct': Decimal('100.00')},
            )

            # ── DMS Document stub (BHR* number only) ──────────────────
            if dms_number and dms_number.upper() not in ('NO', 'N/A', ''):
                doc, doc_created = Document.objects.get_or_create(
                    dms_file_number=dms_number,
                    defaults={
                        'original_filename': f'{dms_number}.pdf',
                        'document_type':     'patta',
                        'status':            'linked',
                        'linked_patta':      patta,
                        'uploaded_by_id':    system_user_id,
                    },
                )
                if doc_created:
                    counts['documents'] += 1
                # Link patta → document if not already linked
                if not patta.document_id:
                    Patta.objects.filter(pk=patta.pk).update(document_id=doc.pk)
                    patta.document_id = doc.pk  # keep in-memory object in sync

        self.stdout.write(
            f'  → plots={counts["plots"]}  pattas={counts["pattas"]}  '
            f'khasras={counts["khasras"]}  docs={counts["documents"]}'
        )
        return counts


class _DryRunRollback(Exception):
    """Sentinel exception used to roll back a dry-run transaction."""
    pass
