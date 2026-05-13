"""
import_rejected_layouts — load the East/West rejected-layout Excel.

The Authority shares rejected-layout decisions as a 2-sheet .xlsx
(one sheet per zone) with this header (in Krutidev-encoded Hindi):

    क्र.सं. | कॉलोनी का नाम | आवेदन की दिनांक | राजस्व ग्राम | खसरा नम्बर | अस्वीकार करने का कारण

We import each row as a Colony with:

    colony_type     = 'rejected_layout'
    zone            = <sheet name, East/West>
    name            = Krutidev → Unicode-converted colony name
    revenue_village = English revenue village string (already Latin)
    khasras_input   = parsed comma-separated khasra numbers
    rejection_reason = Krutidev → Unicode-converted reason
    status          = 'active'   (public views filter on status='active';
                                  rejected layouts are still meant to be
                                  discoverable on the public portal so
                                  citizens can see what was rejected)

Skipped:
  · layout_application_date — by design (column dropped from Colony model)
  · empty rows (the sheets have stray blank rows for visual spacing)

Idempotency: looks up by (name, zone). Existing row is left alone
unless `--update` is passed, in which case rejection_reason +
khasras are refreshed.

Usage:
    docker compose -f docker-compose.prod.yml exec backend \\
        python manage.py import_rejected_layouts \\
            --file /tmp/East-West_Layout_File.xlsx
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from colonies.models import Colony, Khasra
from ._krutidev import krutidev_to_unicode

logger = logging.getLogger(__name__)

# 1-based column indices (matches Excel column letters A..F).
COL_SERIAL          = 1
COL_NAME            = 2
COL_APPLICATION     = 3   # ignored intentionally — column dropped from model
COL_REVENUE_VILLAGE = 4
COL_KHASRA          = 5
COL_REASON          = 6

DATA_FIRST_ROW = 3   # row 1 is sheet title, row 2 is the column header
VALID_ZONES    = {'East', 'West'}


def _norm_text(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


def _parse_khasras(raw) -> list[str]:
    """
    Khasra cells come in three shapes:
      1. Comma-separated:  '178, 179, 180/574'
      2. Concatenated:     '209212'         → ['209212']  (no spaces, keep as-is)
      3. Sometimes mixed:  '768777778681'   → ['768', '777', '778', '681']
         when it's a run of 3-digit numbers with no separator.

    Excel often types these as numbers, losing the separators. We can
    only do so much: if the cell is comma-separated we split cleanly;
    if it's a pure integer with > 3 digits and length is a multiple of
    3 with all parts looking like plausible khasra numbers, we split
    every 3 chars. Otherwise we keep the whole value as a single khasra.
    """
    s = _norm_text(raw).replace('\n', ',').replace(';', ',')
    # Convert ints like 925.0 from openpyxl into plain '925'
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        s = str(int(raw))
    if not s:
        return []
    if ',' in s:
        out, seen = [], set()
        for tok in s.split(','):
            t = re.sub(r'\s+', '', tok)
            if t and t not in seen:
                seen.add(t)
                out.append(t)
        return out
    # Pure-digit concatenation: try splitting every 3 chars only when
    # length is exactly a multiple of 3 between 6 and 21 digits — that's
    # the realistic range for 2–7 khasras squished together.
    if s.isdigit() and 6 <= len(s) <= 21 and len(s) % 3 == 0:
        parts = [s[i:i+3] for i in range(0, len(s), 3)]
        return parts
    return [s]


class Command(BaseCommand):
    help = 'Import rejected colony layouts from the Authority Excel sheet.'

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True, type=str,
                            help='Path to the .xlsx file.')
        parser.add_argument('--update', action='store_true',
                            help='Refresh khasras + rejection_reason on existing rows.')
        parser.add_argument('--dry-run', action='store_true',
                            help='Parse only — report what would happen.')

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl missing — should be in requirements.txt.')

        path = Path(opts['file']).expanduser()
        if not path.exists():
            raise CommandError(f'File not found: {path}')

        wb     = openpyxl.load_workbook(path, data_only=True)
        zones  = [s for s in wb.sheetnames if s in VALID_ZONES]
        if not zones:
            raise CommandError(
                f'No East/West sheets found. Got: {wb.sheetnames}'
            )

        totals = {'seen': 0, 'created': 0, 'updated': 0, 'skipped': 0, 'errored': 0}

        with transaction.atomic():
            for zone in zones:
                ws = wb[zone]
                self.stdout.write(self.style.MIGRATE_HEADING(
                    f'── {zone} zone — {ws.max_row} rows'
                ))
                for r in range(DATA_FIRST_ROW, ws.max_row + 1):
                    raw_serial = ws.cell(r, COL_SERIAL).value
                    raw_name   = ws.cell(r, COL_NAME).value
                    if raw_serial in (None, '') and raw_name in (None, ''):
                        continue  # blank spacer row
                    totals['seen'] += 1
                    try:
                        action = self._process_row(zone, ws, r, opts)
                        totals[action] += 1
                    except Exception as exc:
                        totals['errored'] += 1
                        logger.error('row %s/%s failed: %s', zone, r, exc,
                                     exc_info=True)
                        self.stdout.write(self.style.ERROR(
                            f'  ✗ {zone} r{r}: {exc}'
                        ))

            if opts['dry_run']:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING(
                    '── DRY RUN — rolling back transaction'
                ))

        self.stdout.write(self.style.SUCCESS(
            f'Done. seen={totals["seen"]}  created={totals["created"]}  '
            f'updated={totals["updated"]}  skipped={totals["skipped"]}  '
            f'errored={totals["errored"]}'
        ))

    # ── per-row logic ───────────────────────────────────────────────────────

    def _process_row(self, zone, ws, r, opts):
        krd_name        = _norm_text(ws.cell(r, COL_NAME).value)
        revenue_village = _norm_text(ws.cell(r, COL_REVENUE_VILLAGE).value)
        khasra_raw      = ws.cell(r, COL_KHASRA).value
        krd_reason      = _norm_text(ws.cell(r, COL_REASON).value)

        name_u   = krutidev_to_unicode(krd_name)
        reason_u = krutidev_to_unicode(krd_reason)
        khasras  = _parse_khasras(khasra_raw)

        if not name_u:
            self.stdout.write(self.style.WARNING(
                f'  · {zone} r{r}: blank name, skipped'
            ))
            return 'skipped'

        colony, created = Colony.objects.get_or_create(
            name=name_u, zone=zone,
            defaults={
                'colony_type':       'rejected_layout',
                'status':            'active',
                'revenue_village':   revenue_village,
                'rejection_reason':  reason_u,
            },
        )

        # Sync khasras (always for new rows; on --update for existing rows).
        if created or opts['update']:
            existing = set(colony.khasras.values_list('number', flat=True))
            for k in khasras:
                if k not in existing:
                    Khasra.objects.get_or_create(colony=colony, number=k)
                    existing.add(k)

        if created:
            self.stdout.write(self.style.SUCCESS(
                f'  ✓ {zone} r{r}: created "{name_u}" '
                f'(village={revenue_village or "—"}, {len(khasras)} khasras)'
            ))
            return 'created'

        if opts['update']:
            changed = False
            if colony.rejection_reason != reason_u:
                colony.rejection_reason = reason_u
                changed = True
            if colony.revenue_village != revenue_village and revenue_village:
                colony.revenue_village = revenue_village
                changed = True
            if changed:
                colony.colony_type = 'rejected_layout'
                colony.save(update_fields=[
                    'rejection_reason', 'revenue_village',
                    'colony_type', 'updated_at',
                ])
                self.stdout.write(self.style.NOTICE(
                    f'  ↻ {zone} r{r}: updated "{name_u}"'
                ))
                return 'updated'

        self.stdout.write(f'  · {zone} r{r}: "{name_u}" already exists, skipped')
        return 'skipped'
